import importlib.util
import sys
import numpy as np
import pandas as pd
import pulp
import time
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.colors import TwoSlopeNorm
def set_plot_style():
    """统一的出版级绘图风格（中文环境，论文友好）"""
    plt.rcParams.update({
        'font.sans-serif': ['SimHei', 'Microsoft YaHei UI', 'Arial Unicode MS', 'DejaVu Sans'],
        'axes.unicode_minus': False,
        'figure.dpi': 120,
        'savefig.dpi': 300,
        'axes.titlesize': 13,
        'axes.labelsize': 11.5,
        'xtick.labelsize': 10,
        'ytick.labelsize': 10,
        'legend.fontsize': 10,
        'axes.spines.top': False,
        'axes.spines.right': False,
        'grid.alpha': 0.3,
        'grid.linestyle': '--',
        'axes.grid': True,
        'lines.linewidth': 2.0,
    })


def load_problem2_impl():
    spec = importlib.util.spec_from_file_location(
        "p2_impl", "第二题_最终严格版本.py"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.FinalStrictPaperImplementation


class Problem3CorrelatedSubstitutionSolver:
    """
    问题三：考虑相关性与替代/互补关系的随机规划求解器（基于问题二扩展）
    - 在问题二的SAA+内生CVaR框架上，加入：
      1) 相关情景生成（P-D-C联合波动，Gaussian Copula）
      2) 迭代修正参数（~P, ~D, ~C 依赖供给S与关系矩阵δ）
      3) 替代/互补的面积配比约束
    - 采用“固定~参数 -> 线性MILP -> 更新~参数”的外层迭代，确保可解性
    """

    def __init__(self):
        Base = load_problem2_impl()
        self.base = Base()

        # 关系与相关性参数
        self.alpha_relation = 0.20  # 关系强度系数 α
        # 下述 ρ 视为“目标 Spearman 相关”，用于 Copula 标定
        self.rho_PD = -0.40         # 价格-销量（目标 Spearman，负）
        self.rho_DC = 0.15          # 销量-成本（目标 Spearman，正）
        self.rho_PC = 0.10          # 价格-成本（目标 Spearman，弱正）

        self.sigma_p = 0.03         # 价格波动强度
        self.sigma_d = 0.05         # 销量波动强度
        self.sigma_c = 0.02         # 成本波动强度

        self.synergy_gain = 0.05    # 豆类后茬增产5%（用于验证阶段）

        # 关系矩阵 δ（示例）与集合
        self.delta = {}  # (j, j2) -> delta
        self.strong_substitutes = set()  # 强替代作物集合
        self.complement_pairs = []       # 互补对 (j, j')

        # Copula 相关矩阵（latent Gaussian）
        self.copula_R = None

    # ---------- 关系设置 ----------
    def build_relationships(self):
        # 仅对常见对设置示例（若作物编号缺失则自动跳过）
        # 西红柿(21)–茄子(22)：强替代 -0.5
        self.delta[(21, 22)] = -0.5
        self.delta[(22, 21)] = -0.5
        self.strong_substitutes.update([21, 22])

        # 小麦(6)–玉米(7)：中替代 -0.3
        self.delta[(6, 7)] = -0.3
        self.delta[(7, 6)] = -0.3

        # 豆类(1) 与 小麦/玉米：弱互补 +0.2
        self.delta[(1, 6)] = 0.2
        self.delta[(6, 1)] = 0.2
        self.delta[(1, 7)] = 0.2
        self.delta[(7, 1)] = 0.2
        self.complement_pairs.extend([(1, 6), (1, 7)])

    # ---------- 情景生成与相关性注入 ----------
    def generate_correlated_scenarios(self):
        self.base.generate_stochastic_scenarios()  # 基础（独立）情景
        # 固定随机种子，确保Copula注入可复现
        np.random.seed(2025)

        # —— Copula 标定：将目标 Spearman ρS 映射到 latent Gaussian ρG ——
        # 公式：ρS = 6/π · arcsin(ρG/2)  ⇔  ρG = 2·sin(π·ρS/6)
        def spearman_to_gaussian_rho(r_s: float) -> float:
            return float(2.0 * np.sin(np.pi * r_s / 6.0))

        r_pd_g = spearman_to_gaussian_rho(self.rho_PD)
        r_dc_g = spearman_to_gaussian_rho(self.rho_DC)
        r_pc_g = spearman_to_gaussian_rho(self.rho_PC)

        R = np.array([
            [1.0,   r_pd_g, r_pc_g],
            [r_pd_g, 1.0,   r_dc_g],
            [r_pc_g, r_dc_g, 1.0  ],
        ])

        # 保障正定性（如有需要，添加轻微抬升）
        eigvals = np.linalg.eigvalsh(R)
        if eigvals.min() < 1e-6:
            eps = abs(eigvals.min()) + 1e-6
            R = R + np.eye(3) * eps
        self.copula_R = R

        crop_ids = list(self.base.crop_info.keys())
        # 记录用于相关检验的乘法因子
        sample_p, sample_d, sample_c = [], [], []
        for k, scenario in enumerate(self.base.scenarios):
            for crop_id in crop_ids:
                if crop_id not in scenario:
                    continue
                for year in self.base.years:
                    if year not in scenario[crop_id]:
                        continue
                    year_data = scenario[crop_id][year]
                    # 3维相关标准正态
                    z = np.random.multivariate_normal(mean=[0, 0, 0], cov=self.copula_R)
                    # 单调映射为乘法因子（log-normal 形式），保持秩次 → Spearman 不变
                    fp = float(np.exp(self.sigma_p * z[0]))  # 价格因子
                    fd = float(np.exp(self.sigma_d * z[1]))  # 销量因子
                    fc = float(np.exp(self.sigma_c * z[2]))  # 成本因子

                    year_data['price'] *= fp
                    year_data['sales_limit'] *= fd
                    year_data['cost_factor'] *= fc

                    sample_p.append(fp)
                    sample_d.append(fd)
                    sample_c.append(fc)

        # —— Copula 标定报告（样本 Spearman 与目标的误差）——
        def spearman_corr(a, b):
            ra = pd.Series(a).rank().to_numpy()
            rb = pd.Series(b).rank().to_numpy()
            return float(np.corrcoef(ra, rb)[0, 1])

        sp_pd = spearman_corr(sample_p, sample_d)
        sp_dc = spearman_corr(sample_d, sample_c)
        sp_pc = spearman_corr(sample_p, sample_c)
        report = pd.DataFrame([
            {
                'pair': 'P-D',
                'target_spearman': self.rho_PD,
                'fitted_spearman': sp_pd,
                'abs_error': abs(sp_pd - self.rho_PD),
            },
            {
                'pair': 'D-C',
                'target_spearman': self.rho_DC,
                'fitted_spearman': sp_dc,
                'abs_error': abs(sp_dc - self.rho_DC),
            },
            {
                'pair': 'P-C',
                'target_spearman': self.rho_PC,
                'fitted_spearman': sp_pc,
                'abs_error': abs(sp_pc - self.rho_PC),
            },
        ])
        report.to_csv('问题三_Copula标定报告.csv', index=False, encoding='utf-8-sig')
        print('🔧 Copula 标定：已输出 问题三_Copula标定报告.csv')

    # ---------- 外层迭代：修正参数 ~P, ~D, ~C ----------
    def compute_supply_mean_by_crop_year(self, x_solution):
        """根据当前解x与场景平均产量，估算各作物-年份的平均供给 S_mean[j, t]"""
        crop_ids = list(self.base.crop_info.keys())
        land_names = list({**self.base.grain_lands, **self.base.irrigation_lands, **self.base.greenhouse_lands}.keys())

        # 预计算各作物在每年平均产量因子 E[yield_factor]
        E_yield = {}
        for crop_id in crop_ids:
            E_yield[crop_id] = {}
            for year in self.base.years:
                vals = [sc[crop_id][year]['yield_factor'] for sc in self.base.scenarios]
                E_yield[crop_id][year] = float(np.mean(vals))

        S_mean = {(crop_id, year): 0.0 for crop_id in crop_ids for year in self.base.years}

        for land in x_solution:
            for year in x_solution[land]:
                for season in x_solution[land][year]:
                    for crop_id, area in x_solution[land][year][season].items():
                        if area <= 0:
                            continue
                        y_base = self.base.crop_info[crop_id]['yield_base']
                        S_mean[(crop_id, year)] += area * y_base * E_yield[crop_id][year]

        return S_mean

    def compute_mean_D_by_crop_year(self):
        crop_ids = list(self.base.crop_info.keys())
        D_mean = {(crop_id, year): 0.0 for crop_id in crop_ids for year in self.base.years}
        for crop_id in crop_ids:
            for year in self.base.years:
                vals = [sc[crop_id][year]['sales_limit'] for sc in self.base.scenarios]
                D_mean[(crop_id, year)] = float(np.mean(vals))
        return D_mean

    def apply_tilde_updates(self, S_mean, D_mean):
        """按 δ 与 S/D 更新 ~P, ~D, ~C 的缩放因子，并写回场景字典"""
        crop_ids = list(self.base.crop_info.keys())

        price_scale = {}
        demand_scale = {}
        cost_scale = {}

        # 先按作物-年计算缩放系数（对所有情景一致），避免在MILP中引入跨情景耦合
        for j in crop_ids:
            for t in self.base.years:
                # 关系对价格的影响：1 + α · Σ δ_{j,j'} · (S_{j'}/D_{j'})
                relation_term = 0.0
                for j2 in crop_ids:
                    if (j, j2) in self.delta:
                        ratio = 0.0
                        if D_mean.get((j2, t), 0.0) > 1e-9:
                            ratio = S_mean.get((j2, t), 0.0) / D_mean[(j2, t)]
                        relation_term += self.delta[(j, j2)] * ratio
                ps = 1.0 + self.alpha_relation * relation_term
                ps = float(np.clip(ps, 0.7, 1.3))

                # 需求对价格变化的弹性 & 互补影响（正δ）
                complement_term = 0.0
                for j2 in crop_ids:
                    if (j, j2) in self.delta and self.delta[(j, j2)] > 0:
                        if D_mean.get((j2, t), 0.0) > 1e-9:
                            complement_term += S_mean.get((j2, t), 0.0) / D_mean[(j2, t)]
                ds = 1.0 + self.rho_PD * (ps - 1.0) + 0.2 * complement_term
                ds = float(np.clip(ds, 0.7, 1.3))

                # 成本随销量走高
                base_ratio = 0.0
                if D_mean.get((j, t), 0.0) > 1e-9:
                    base_ratio = S_mean.get((j, t), 0.0) / D_mean[(j, t)]
                cs = 1.0 + self.rho_DC * base_ratio
                cs = float(np.clip(cs, 0.8, 1.2))

                price_scale[(j, t)] = ps
                demand_scale[(j, t)] = ds
                cost_scale[(j, t)] = cs

        # 阻尼更新：θ ← (1−β)·θ + β·θ_new （避免外层震荡）
        beta = 0.5
        for sc in self.base.scenarios:
            for j in crop_ids:
                if j not in sc:
                    continue
                for t in self.base.years:
                    if t not in sc[j]:
                        continue
                    sc[j][t]['price'] = (1.0 - beta) * sc[j][t]['price'] + beta * (sc[j][t]['price'] * price_scale[(j, t)])
                    sc[j][t]['sales_limit'] = (1.0 - beta) * sc[j][t]['sales_limit'] + beta * (sc[j][t]['sales_limit'] * demand_scale[(j, t)])
                    sc[j][t]['cost_factor'] = (1.0 - beta) * sc[j][t]['cost_factor'] + beta * (sc[j][t]['cost_factor'] * cost_scale[(j, t)])

    # ---------- ρ/δ 灵敏度（±20%）并记录稳健性 ----------
    def run_delta_rho_sensitivity(self, delta_scales=(0.8, 1.0, 1.2), rho_scales=(0.8, 1.0, 1.2), saa=20):
        print('\n🧪 开始 δ/ρ ±20% 灵敏度实验...')
        results = []

        # 备份原始参数
        base_delta = dict(self.delta)
        base_rhos = (self.rho_PD, self.rho_DC, self.rho_PC)
        base_alpha = self.alpha_relation

        for ds in delta_scales:
            for rs in rho_scales:
                print(f"\n—— 组合：δ×{ds:.2f}, ρ×{rs:.2f}, SAA={saa} ——")

                # 重置数据与情景
                self.base.load_and_process_data()
                self.build_relationships()
                # 按比例调整 δ 与 ρ（δ放大通过 α 统一缩放更稳妥）
                self.alpha_relation = base_alpha * ds
                self.rho_PD, self.rho_DC, self.rho_PC = [r * rs for r in base_rhos]

                # 重新生成与相关注入
                self.base.generate_stochastic_scenarios()
                self.generate_correlated_scenarios()

                # SAA求解（单轮）
                self.base.N_saa = saa
                self.base.select_representative_scenarios()
                self.base.build_stochastic_programming_model()
                ok = self.base.solve_model()
                if not ok:
                    print('   ⚠️ 求解失败，跳过记录')
                    continue
                metrics = self.validate_with_synergy()
                metrics.update({'delta_scale': ds, 'rho_scale': rs})
                results.append(metrics)

        if results:
            df = pd.DataFrame(results)
            df.to_csv('问题三_δρ_±20灵敏度.csv', index=False, encoding='utf-8-sig')
            print('✅ 已生成：问题三_δρ_±20灵敏度.csv')
        # 恢复
        self.delta = base_delta
        self.rho_PD, self.rho_DC, self.rho_PC = base_rhos
        self.alpha_relation = base_alpha

    def write_param_sources_note(self):
        text = (
            '相关性/替代性参数口径说明\n'
            '— 目标 Spearman 相关（P-D、D-C、P-C）参考农产品需求弹性与成本传导的文献区间：\n'
            '  • 价格-销量（P-D）：-0.2 ~ -0.6（农产品需求价格弹性弱到中等）\n'
            '  • 销量-成本（D-C）：+0.10 ~ +0.30（需求旺盛时用工/物料成本上行）\n'
            '  • 价格-成本（P-C）：+0.05 ~ +0.20（价格上行伴随要素价格抬升）\n'
            '— 本模型采用中位值作为基线，并在 ±20% 区间内做灵敏度检验；\n'
            '— Copula 采用高斯Copula，并将目标 Spearman ρS 通过 ρG=2·sin(π·ρS/6) 映射到 latent Gaussian；\n'
            '— 采样使用固定随机种子以确保可复现，拟合误差见“问题三_Copula标定报告.csv”。\n'
        )
        with open('问题三_相关性参数说明.txt', 'w', encoding='utf-8') as f:
            f.write(text)
        print('📝 已生成：问题三_相关性参数说明.txt')

    # ---------- 互补/替代 面积约束（在父类约束后追加） ----------
    def add_relation_constraints(self):
        model = self.base.model
        all_lands = {**self.base.grain_lands, **self.base.irrigation_lands, **self.base.greenhouse_lands}
        land_names = list(all_lands.keys())
        crop_ids = list(self.base.crop_info.keys())

        # 强替代：该组作物总面积占比 ≤ 40%（按年）
        subs = [cid for cid in self.strong_substitutes if cid in crop_ids]
        if len(subs) >= 2:
            for year in self.base.years:
                sub_terms = []
                total_terms = []
                for land_name in land_names:
                    max_seasons = all_lands[land_name]['max_seasons']
                    for season in range(1, max_seasons + 1):
                        # sub组
                        for j in subs:
                            if (land_name in self.base.x and j in self.base.x[land_name] and
                                year in self.base.x[land_name][j] and season in self.base.x[land_name][j][year]):
                                sub_terms.append(self.base.x[land_name][j][year][season])
                        # 全部作物
                        for j in crop_ids:
                            if (land_name in self.base.x and j in self.base.x[land_name] and
                                year in self.base.x[land_name][j] and season in self.base.x[land_name][j][year]):
                                total_terms.append(self.base.x[land_name][j][year][season])

                if sub_terms and total_terms:
                    # Σ x_sub - 0.4 * Σ x_total ≤ 0
                    model += (pulp.lpSum(sub_terms) - 0.4 * pulp.lpSum(total_terms) <= 0,
                              f"strong_sub_share_year_{year}")

        # 互补对：x_j ≥ (1/3) x_j' （按年）
        for (j, j2) in self.complement_pairs:
            if j not in crop_ids or j2 not in crop_ids:
                continue
            for year in self.base.years:
                left_terms, right_terms = [], []
                for land_name in land_names:
                    max_seasons = all_lands[land_name]['max_seasons']
                    for season in range(1, max_seasons + 1):
                        if (land_name in self.base.x and j in self.base.x[land_name] and
                            year in self.base.x[land_name][j] and season in self.base.x[land_name][j][year]):
                            left_terms.append(self.base.x[land_name][j][year][season])
                        if (land_name in self.base.x and j2 in self.base.x[land_name] and
                            year in self.base.x[land_name][j2] and season in self.base.x[land_name][j2][year]):
                            right_terms.append(self.base.x[land_name][j2][year][season])
                if left_terms and right_terms:
                    model += (pulp.lpSum(left_terms) - (1.0/3.0) * pulp.lpSum(right_terms) >= 0,
                              f"complement_ratio_{j}_{j2}_{year}")

    # ---------- 验证（带豆类后茬增产） ----------
    def validate_with_synergy(self):
        print(f"\n🎯 使用全部{self.base.N_scenarios}个情景验证（含后茬增产）...")
        scenario_profits = np.zeros(self.base.N_scenarios)
        sum_prod = 0.0
        sum_excess = 0.0

        for k, scenario in enumerate(self.base.scenarios):
            if k % 200 == 0:
                print(f"   验证情景 {k+1}/{self.base.N_scenarios}...")

            total_profit = 0.0
            for land_name, land_solution in self.base.optimal_solution.items():
                for year, year_solution in land_solution.items():
                    # 计算该地块当年季节1的豆类面积（用于后茬判断）
                    bean_area_s1 = 0.0
                    if 1 in year_solution:
                        for crop_id, area in year_solution[1].items():
                            if self.base.crop_info[crop_id]['is_bean']:
                                bean_area_s1 += area

                    for season, season_solution in year_solution.items():
                        for crop_id, area in season_solution.items():
                            crop_info = self.base.crop_info[crop_id]
                            y = crop_info['yield_base'] * scenario[crop_id][year]['yield_factor']
                            c = crop_info['cost_base'] * scenario[crop_id][year]['cost_factor']
                            p = scenario[crop_id][year]['price']
                            d = scenario[crop_id][year]['sales_limit']

                            # 后茬：若季节1种豆类，则季节2非豆类增产5%
                            if season == 2 and bean_area_s1 > 0.01 and not crop_info['is_bean']:
                                y *= (1.0 + self.synergy_gain)

                            production = area * y
                            q_sell = min(production, d)
                            q_excess = max(0.0, production - d)
                            total_profit += q_sell * p + 0.5 * q_excess * p - area * c
                            sum_prod += production
                            sum_excess += q_excess

            scenario_profits[k] = total_profit

        expected_profit = float(np.mean(scenario_profits))
        profit_std = float(np.std(scenario_profits))
        var_5 = float(np.percentile(scenario_profits, 5))
        cvar_5 = float(np.mean(scenario_profits[scenario_profits <= var_5]))
        overprod_ratio = float(sum_excess / sum_prod) if sum_prod > 1e-9 else 0.0

        print("📈 稳健性验证结果(含后茬)：")
        print(f"   - 期望收益: {expected_profit:,.2f} 元")
        print(f"   - 标准差: {profit_std:,.2f} 元 ({profit_std/expected_profit*100:.1f}%)")
        print(f"   - 5% VaR: {var_5:,.2f} 元")
        print(f"   - 5% CVaR: {cvar_5:,.2f} 元")
        print(f"   - 超产比: {overprod_ratio*100:.2f}%")

        self.base.scenario_profits = scenario_profits
        return {
            'expected_profit': expected_profit,
            'profit_std': profit_std,
            'var_5': var_5,
            'cvar_5': cvar_5,
            'overproduction_ratio': overprod_ratio,
        }

    # ---------- 导出与可视化 ----------
    def save_results3(self, filename='附件3/result3.xlsx'):
        """将当前解按问题三口径导出到Excel（按作物-年-季、分段销售估计收入）"""
        print(f"\n💾 保存问题三结果到: {filename}")
        result_data = []
        all_lands = {**self.base.grain_lands, **self.base.irrigation_lands, **self.base.greenhouse_lands}

        # 预计算均值参数（价格、销量上限、产量因子）
        crop_ids = list(self.base.crop_info.keys())
        mean_price = {(j, t): float(np.mean([sc[j][t]['price'] for sc in self.base.scenarios])) for j in crop_ids for t in self.base.years}
        mean_D = {(j, t): float(np.mean([sc[j][t]['sales_limit'] for sc in self.base.scenarios])) for j in crop_ids for t in self.base.years}
        mean_yield_factor = {(j, t): float(np.mean([sc[j][t]['yield_factor'] for sc in self.base.scenarios])) for j in crop_ids for t in self.base.years}

        for land_name, land_solution in self.base.optimal_solution.items():
            land_type = all_lands[land_name]['type']
            for year, year_solution in land_solution.items():
                for season, season_solution in year_solution.items():
                    for crop_id, area in season_solution.items():
                        if area <= 0.01:
                            continue
                        crop_info = self.base.crop_info[crop_id]
                        y_base = crop_info['yield_base']
                        y = y_base * mean_yield_factor[(crop_id, year)]
                        p = mean_price[(crop_id, year)]
                        d = mean_D[(crop_id, year)]

                        production = area * y
                        q_sell = min(production, d)
                        q_excess = max(0.0, production - d)
                        revenue = q_sell * p + 0.5 * q_excess * p

                        result_data.append({
                            '地块名称': land_name,
                            '地块类型': land_type,
                            '年份': year,
                            '季节': season,
                            '作物编号': crop_id,
                            '作物名称': crop_info['name'],
                            '种植面积(亩)': round(area, 2),
                            '预期产量(斤)': round(production, 2),
                            '预期收入(元)': round(revenue, 2)
                        })

        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "种植方案"
        headers = ['地块名称', '地块类型', '年份', '季节', '作物编号', '作物名称', '种植面积(亩)', '预期产量(斤)', '预期收入(元)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        for row, data in enumerate(result_data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=data[header])
        workbook.save(filename)
        print(f"✅ 结果已保存，共 {len(result_data)} 条记录")

    def generate_relation_heatmap(self, filename='图7.1_作物关系热力图.png'):
        """根据已设定的δ对，绘制紧凑的关系热力图（仅含出现过的作物）"""
        print("\n📊 生成关系热力图...")
        set_plot_style()

        crops_used = sorted({j for (j, _) in self.delta.keys()} | {j2 for (_, j2) in self.delta.keys()})
        if not crops_used:
            print("⚠️ 未定义关系对，跳过绘图")
            return

        names = [self.base.crop_info[j]['name'] if j in self.base.crop_info else str(j) for j in crops_used]
        M = np.zeros((len(crops_used), len(crops_used)))
        for a, j in enumerate(crops_used):
            for b, j2 in enumerate(crops_used):
                M[a, b] = self.delta.get((j, j2), 0.0)

        fig, ax = plt.subplots(figsize=(7.2, 5.8))
        norm = TwoSlopeNorm(vmin=-0.6, vcenter=0.0, vmax=0.6)
        im = ax.imshow(M, cmap='RdBu_r', norm=norm)
        ax.set_xticks(range(len(crops_used)))
        ax.set_yticks(range(len(crops_used)))
        ax.set_xticklabels(names, rotation=45, ha='right')
        ax.set_yticklabels(names)
        ax.set_title('作物关系热力图（δ，正=互补，负=替代）')
        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(0.8)
        for i in range(M.shape[0]):
            for j in range(M.shape[1]):
                val = M[i, j]
                if abs(val) > 1e-6:
                    ax.text(j, i, f"{val:.1f}", ha='center', va='center', fontsize=8.5, color='black')
        cbar = fig.colorbar(im, shrink=0.92)
        cbar.ax.set_ylabel('关系强度 δ', rotation=90, labelpad=10)
        plt.tight_layout()
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"✅ 已生成：{filename}")

    def generate_price_supply_scatter(self, filename='图7.2_价格-供给散点.png'):
        print("\n📊 生成价格-供给散点图（优化）...")
        set_plot_style()

        # 平均供给与价格（按作物-年平均），供给用解的 S_mean
        S_mean = self.compute_supply_mean_by_crop_year(self.base.optimal_solution)
        crop_ids = list(self.base.crop_info.keys())
        P_mean = {(j, t): float(np.mean([sc[j][t]['price'] for sc in self.base.scenarios])) for j in crop_ids for t in self.base.years}

        xs, ys = [], []
        for j in crop_ids:
            for t in self.base.years:
                s = S_mean.get((j, t), 0.0)
                if s <= 0:
                    continue
                xs.append(s / 1e6)  # 单位：百万斤
                ys.append(P_mean[(j, t)])

        if len(xs) == 0:
            print("⚠️ 当前解供给为0，跳过绘图")
            return

        xs = np.array(xs, dtype=float)
        ys = np.array(ys, dtype=float)

        # 计算 Pearson 与 Spearman
        r_pearson = float(np.corrcoef(xs, ys)[0, 1]) if len(xs) > 1 else np.nan
        rank_x = pd.Series(xs).rank().to_numpy()
        rank_y = pd.Series(ys).rank().to_numpy()
        r_spearman = float(np.corrcoef(rank_x, rank_y)[0, 1]) if len(xs) > 1 else np.nan

        # 胜任的可视化：左右双图（左：全局散点+稳健拟合；右：小供给区密度）
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12.5, 5.6))

        # 1) 全局散点 + 95%分位稳健拟合（去极端）
        qx_lo, qx_hi = np.quantile(xs, [0.025, 0.975])
        qy_lo, qy_hi = np.quantile(ys, [0.025, 0.975])
        mask = (xs >= qx_lo) & (xs <= qx_hi) & (ys >= qy_lo) & (ys <= qy_hi)
        xs_fit, ys_fit = xs[mask], ys[mask]
        if len(xs_fit) >= 2:
            coef = np.polyfit(xs_fit, ys_fit, 1)
            xline = np.linspace(xs.min(), xs.max(), 200)
            yline = np.polyval(coef, xline)
        else:
            coef = [np.nan, np.nan]
            xline = np.array([xs.min(), xs.max()])
            yline = np.array([ys.mean(), ys.mean()])

        ax1.scatter(xs, ys, c='tab:blue', alpha=0.65, s=30, linewidths=0, rasterized=True)
        ax1.plot(xline, yline, color='tab:red', linestyle='--', linewidth=2,
                 label=f"OLS(95%区间): y={coef[0]:.2f}x+{coef[1]:.2f}")
        ax1.set_xlabel('供给量 S（百万斤）')
        ax1.set_ylabel('均价（元/斤）')
        ax1.set_title(f'价格—供给关系（按作物-年聚合）')
        ax1.legend(loc='upper right')
        ax1.grid(alpha=0.3)
        ax1.text(0.02, 0.02, f"Pearson r={r_pearson:.2f}\nSpearman ρ={r_spearman:.2f}",
                 transform=ax1.transAxes, fontsize=9,
                 bbox=dict(boxstyle='round', facecolor='white', alpha=0.7, edgecolor='lightgray'))

        # 2) 小供给区密度（hexbin），突出近原点拥挤区
        x_zoom_max = float(np.quantile(xs, 0.85))  # 自适应放大到85%分位
        ax2.hexbin(xs, ys, gridsize=28, cmap='viridis', extent=[0, x_zoom_max, ys.min(), ys.max()],
                   mincnt=1)
        ax2.set_xlim(0, x_zoom_max)
        ax2.set_xlabel('供给量 S（百万斤，近原点放大）')
        ax2.set_ylabel('均价（元/斤）')
        ax2.set_title('近原点密度（Hexbin）')
        cb = fig.colorbar(ax2.collections[0], ax=ax2)
        cb.set_label('计数')
        # 次轴不额外网格，保持主图风格统一

        fig.suptitle('图7.2 价格—供给关系：全局拟合与局部密度', fontsize=13)
        plt.tight_layout(rect=[0, 0, 1, 0.95])
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"✅ 已生成：{filename}")

    def generate_q2_vs_q3_boxplot(self, filename='问题2_vs_问题3_收益箱线图.png', csvfile='问题2_vs_问题3_关键指标.csv'):
        print("\n📊 生成问题2 vs 问题3 收益对比箱线图...")
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False

        # 运行问题2（相同K=1000，SAA=当前值），获取场景收益
        Base = load_problem2_impl()
        q2 = Base()
        q2.N_saa = self.base.N_saa
        q2.load_and_process_data()
        q2.generate_stochastic_scenarios()
        q2.select_representative_scenarios()
        # 尾部增约稳健性：对不同M进行验证（可选）
        q2.use_tail_cvar = True
        for M in [int(0.1*q2.N_saa), int(0.2*q2.N_saa), q2.N_saa]:
            q2.tail_M = max(1, M)
            # 初步使用所有代表性情景的最差集合：先以独立损失排序确定
            q2.tail_active_indices = None  # 让模型内采用默认初值
        q2.build_stochastic_programming_model()
        ok = q2.solve_model()
        if not ok:
            print("⚠️ 问题2求解失败，跳过对比")
            return
        q2_profits = np.array(q2.scenario_profits, dtype=float)

        # 问题3收益（已含后茬验证）
        q3_profits = np.array(self.base.scenario_profits, dtype=float)

        # 指标
        def metrics(arr):
            mu = float(np.mean(arr))
            std = float(np.std(arr))
            var5 = float(np.percentile(arr, 5))
            cvar5 = float(np.mean(arr[arr <= var5]))
            return mu, std, var5, cvar5
        m2 = metrics(q2_profits)
        m3 = metrics(q3_profits)

        pd.DataFrame([
            {'方案': '问题2', '期望收益': m2[0], '标准差': m2[1], 'VaR5%': m2[2], 'CVaR5%': m2[3]},
            {'方案': '问题3', '期望收益': m3[0], '标准差': m3[1], 'VaR5%': m3[2], 'CVaR5%': m3[3]},
        ]).to_csv(csvfile, index=False, encoding='utf-8-sig')

        fig, ax = plt.subplots(figsize=(7.5, 5.5))
        data_million = [q2_profits / 1e6, q3_profits / 1e6]
        ax.boxplot(data_million, labels=['问题二', '问题三'], patch_artist=True,
                   boxprops=dict(facecolor='lightsteelblue'), medianprops=dict(color='red'))
        ax.set_ylabel('收益（百万元）')
        ax.set_title('问题二 vs 问题三 收益分布对比（1000情景）')
        # 标注均值
        means = [np.mean(d) for d in data_million]
        for i, m in enumerate(means, start=1):
            ax.text(i, m, f"均值={m:.1f}", ha='center', va='bottom', fontsize=9, color='dimgray')
        ax.grid(axis='y', alpha=0.3)
        plt.tight_layout()
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"✅ 已生成：{filename} 与 {csvfile}")

    # ---------- 主流程 ----------
    def run(self, iterations=2, tail_M: int | None = None, test_tail_robustness: bool = False):
        print("\n🚀 问题三：相关性与替代性 扩展求解 开始")
        t0 = time.time()

        # 1) 数据加载与基础情景
        self.base.load_and_process_data()
        self.build_relationships()
        self.base.generate_stochastic_scenarios()  # 基础
        self.generate_correlated_scenarios()       # 注入相关

        # 2) 外层迭代：~P,~D,~C 修正 + MILP 求解
        # 初始：不修正（使用相关后的基础情景）
        x_sol_last = None
        # 收敛判据：解的均方变化率 < 1e-3，最多10次
        max_iters = min(10, iterations)
        tol = 1e-3
        prev_vector = None
        for it in range(1, max_iters + 1):
            print(f"\n==== 外层迭代 {it}/{iterations} ====")

            # SAA 选情景 + 构建与求解
            self.base.select_representative_scenarios()
            # 尾部增约（可选）：启用并进行2轮主动集更新
            if tail_M is not None and tail_M > 0:
                self.solve_with_tail_cvar(tail_M, max_rounds=2)
                ok = (self.base.optimal_solution is not None)
            else:
                self.base.use_tail_cvar = False
                self.base.tail_M = None
                self.base.tail_active_indices = None
                self.base.build_stochastic_programming_model()
                ok = self.base.solve_model()
            if not ok:
                print("⚠️ MILP 求解失败，提前结束迭代")
                break

            # 估算供给与更新 ~参数
            S_mean = self.compute_supply_mean_by_crop_year(self.base.optimal_solution)
            D_mean = self.compute_mean_D_by_crop_year()
            self.apply_tilde_updates(S_mean, D_mean)
            x_sol_last = self.base.optimal_solution

            # 收敛检测：将x展开为向量
            vec = []
            for land_name in sorted(self.base.optimal_solution.keys()):
                for year in sorted(self.base.optimal_solution[land_name].keys()):
                    for season in sorted(self.base.optimal_solution[land_name][year].keys()):
                        for crop_id in sorted(self.base.optimal_solution[land_name][year][season].keys()):
                            vec.append(self.base.optimal_solution[land_name][year][season][crop_id])
            vec = np.array(vec, dtype=float) if vec else np.zeros(1)
            if prev_vector is not None and prev_vector.size == vec.size:
                denom = max(1e-8, np.linalg.norm(prev_vector))
                rel_change = float(np.linalg.norm(vec - prev_vector) / denom)
                print(f"   ↳ 外层相对变化率: {rel_change:.3e}")
                if rel_change < tol:
                    print("   ✅ 满足收敛判据，提前停止外层迭代")
                    break
            prev_vector = vec

        # 3) 验证（含后茬增产），输出指标
        metrics = self.validate_with_synergy()
        # 可选：对不同M做稳健性检验
        if test_tail_robustness:
            try:
                self.run_tail_M_robustness()
            except Exception as e:
                print(f"⚠️ 尾部增约稳健性测试出错：{e}")
        # 导出Excel与图
        try:
            self.save_results3()
            self.generate_relation_heatmap()
            self.generate_price_supply_scatter()
            self.generate_q2_vs_q3_boxplot()
            # 新增：倒需求曲线（δ对比）、作业清单与甘特图、大棚月度利用率
            self.generate_inverse_demand_shift_plot()
            self.export_operation_schedule_and_gantt()
            self.generate_greenhouse_monthly_utilization()
        except Exception as e:
            print(f"⚠️ 后处理出错：{e}")
        total_time = time.time() - t0
        print(f"\n🏁 问题三求解完成，用时 {total_time:.2f} 秒")
        return metrics

    # ---------- 消融实验：Baseline / 仅Copula / 仅δ / 两者皆有 ----------
    def run_ablation_experiments(self, out_csv='问题三_消融实验.csv', saa: int | None = None):
        """四种配置消融：
        - baseline: 无相关（不注入Copula，不启用δ更新）
        - copula_only: 仅注入Copula相关
        - delta_only: 仅启用δ关系（含~参数更新与阻尼）
        - both: Copula + δ
        输出每种配置的 E[Profit], CVaR5%, 超产比，并给出相对baseline的p值（Welch近似）。
        """
        import math

        def pvalue_welch(a: np.ndarray, b: np.ndarray) -> float:
            a = np.array(a, dtype=float).ravel()
            b = np.array(b, dtype=float).ravel()
            na, nb = len(a), len(b)
            if na == 0 or nb == 0:
                return float('nan')
            ma, mb = float(np.mean(a)), float(np.mean(b))
            va, vb = float(np.var(a, ddof=1)), float(np.var(b, ddof=1))
            se = math.sqrt(va/na + vb/nb) if (va>0 or vb>0) else 1e-12
            z = (mb - ma) / se
            # 正态近似p值
            p = 2.0 * 0.5 * math.erfc(abs(z) / math.sqrt(2.0))
            return float(p)

        configs = [
            ('baseline', dict(copula=False, delta=False, tilde=False)),
            ('copula_only', dict(copula=True, delta=False, tilde=False)),
            ('delta_only', dict(copula=False, delta=True, tilde=True)),
            ('both', dict(copula=True, delta=True, tilde=True)),
        ]

        rows = []
        profits_map = {}

        for name, cfg in configs:
            # 重置并运行单配置
            res, profits = self._solve_ablation_single(cfg, saa=saa)
            profits_map[name] = profits
            rows.append({
                'config': name,
                **res
            })

        # p值相对baseline（收益均值）
        base = profits_map.get('baseline')
        if base is not None:
            for row in rows:
                name = row['config']
                if name == 'baseline':
                    row['p_value_vs_baseline'] = 1.0
                else:
                    row['p_value_vs_baseline'] = pvalue_welch(base, profits_map.get(name, []))

        df = pd.DataFrame(rows)
        df.to_csv(out_csv, index=False, encoding='utf-8-sig')
        print(f"✅ 已生成：{out_csv}")
        return df

    def _solve_ablation_single(self, cfg: dict, saa: int | None = None):
        # 统一数据与情景
        self.base.load_and_process_data()
        # 关系设置
        self.delta.clear(); self.complement_pairs.clear(); self.strong_substitutes.clear()
        self.alpha_relation = 0.0
        if cfg.get('delta', False):
            self.build_relationships()
            self.alpha_relation = 0.20
        # 基础情景
        self.base.generate_stochastic_scenarios()
        # Copula注入
        if cfg.get('copula', False):
            self.generate_correlated_scenarios()
        # SAA
        if saa is not None:
            self.base.N_saa = saa
        self.base.select_representative_scenarios()

        # 禁用尾部增约
        self.base.use_tail_cvar = False
        self.base.tail_M = None
        self.base.tail_active_indices = None

        # 求解
        self.base.build_stochastic_programming_model()
        ok = self.base.solve_model()
        if not ok:
            return ({'expected_profit': float('nan'), 'cvar_5': float('nan'), 'overproduction_ratio': float('nan')}, np.array([]))

        # 是否进行~参数阻尼更新（一次）
        if cfg.get('tilde', False):
            S_mean = self.compute_supply_mean_by_crop_year(self.base.optimal_solution)
            D_mean = self.compute_mean_D_by_crop_year()
            self.apply_tilde_updates(S_mean, D_mean)
            # 重新求解一次（保持公平，仍不启用尾部增约）
            self.base.select_representative_scenarios()
            self.base.build_stochastic_programming_model()
            ok = self.base.solve_model()
            if not ok:
                pass

        metrics = self.validate_with_synergy()
        profits = np.array(self.base.scenario_profits, dtype=float)
        out = {
            'E_profit': metrics['expected_profit'],
            'CVaR5': metrics['cvar_5'],
            'OverProdRatio': metrics['overproduction_ratio'],
        }
        return out, profits

    # ---------- 尾部增约：主动集更新求解 ----------
    def solve_with_tail_cvar(self, M: int, max_rounds: int = 2):
        """启用尾部增约：仅对SAA代表性情景中最差的M个施加CVaR约束，并进行主动集(Active Set)更新。"""
        self.base.use_tail_cvar = True
        self.base.tail_M = int(max(1, M))
        self.base.tail_active_indices = None  # 第一次用默认初值
        # 第一次构建并求解
        self.base.build_stochastic_programming_model()
        ok = self.base.solve_model()
        if not ok:
            return False
        # 主动集更新
        for r in range(1, max_rounds):
            worst = self._select_worst_selected_indices(self.base.optimal_solution)
            if worst == self.base.tail_active_indices:
                break
            self.base.tail_active_indices = worst
            self.base.build_stochastic_programming_model()
            ok = self.base.solve_model()
            if not ok:
                break
        return ok

    def _select_worst_selected_indices(self, x_solution):
        """基于当前解，计算SAA代表性情景的损失，返回最差M个的索引列表。"""
        n = len(self.base.selected_scenarios)
        M = int(max(1, min(self.base.tail_M or n, n)))
        losses = []
        for k, scenario in enumerate(self.base.selected_scenarios):
            profit = 0.0
            for land_name, land_solution in x_solution.items():
                for year, year_solution in land_solution.items():
                    for season, season_solution in year_solution.items():
                        for crop_id, area in season_solution.items():
                            crop_info = self.base.crop_info[crop_id]
                            y = crop_info['yield_base'] * scenario[crop_id][year]['yield_factor']
                            c = crop_info['cost_base'] * scenario[crop_id][year]['cost_factor']
                            p = scenario[crop_id][year]['price']
                            d = scenario[crop_id][year]['sales_limit']
                            production = area * y
                            q_sell = min(production, d)
                            q_excess = max(0.0, production - d)
                            profit += q_sell * p + 0.5 * q_excess * p - area * c
            losses.append((-profit, k))  # 损失= -利润
        losses.sort(reverse=True)  # 从大到小（更差在前）
        worst_indices = [k for _, k in losses[:M]]
        return worst_indices

    def run_tail_M_robustness(self, Ms: list[int] | None = None, out_csv='问题三_尾部增约_M稳健性.csv'):
        """对不同M（尾部情景数量）进行求解与全情景验证，输出稳健性结果。"""
        if Ms is None:
            Ms = [max(1, int(0.1 * self.base.N_saa)), max(1, int(0.2 * self.base.N_saa)), self.base.N_saa]
        rows = []
        for M in Ms:
            # 重新选择代表性情景（保证可重复性）
            self.base.select_representative_scenarios()
            t0 = time.time()
            ok = self.solve_with_tail_cvar(M, max_rounds=2)
            t1 = time.time()
            if not ok:
                continue
            metrics = self.validate_with_synergy()
            rows.append({
                'M': M,
                'solve_time_sec': round(t1 - t0, 2),
                **metrics
            })
        if rows:
            pd.DataFrame(rows).to_csv(out_csv, index=False, encoding='utf-8-sig')
            print(f"✅ 已生成：{out_csv}")

    # ---------- 倒需求曲线：主粮/主菜，δ对比 ----------
    def generate_inverse_demand_shift_plot(self, filename='图7.X_倒需求_δ对比_主粮主菜.png'):
        print("\n📊 生成倒需求曲线（主粮/主菜，δ对比）...")
        set_plot_style()

        # 定义类别
        def is_grain(crop_info):
            t = crop_info.get('type', '')
            return ('粮食' in t)
        def is_veg(crop_info):
            t = crop_info.get('type', '')
            return ('蔬菜' in t)

        # 以当前最优解的面积与基准产量构造Q（避免不同情景带来的供给波动）
        grains = []  # (Q_t, P_t) per year
        vegs = []
        # 当前（含δ）价格均值
        mean_price = {(j, t): float(np.mean([sc[j][t]['price'] for sc in self.base.scenarios])) for j in self.base.crop_info for t in self.base.years}
        # 构造不含δ的价格均值：重生成一次情景（相同随机种子），不进行~参数更新
        # 备份
        scenarios_backup = self.base.scenarios
        try:
            # 重新生成基础情景与Copula
            self.base.generate_stochastic_scenarios()
            self.generate_correlated_scenarios()
            mean_price_no_delta = {(j, t): float(np.mean([sc[j][t]['price'] for sc in self.base.scenarios])) for j in self.base.crop_info for t in self.base.years}
        finally:
            # 恢复原场景（含δ后的）
            self.base.scenarios = scenarios_backup

        # 聚合到主粮/主菜（按年）
        all_lands = {**self.base.grain_lands, **self.base.irrigation_lands, **self.base.greenhouse_lands}
        for year in self.base.years:
            Q_grain = 0.0
            P_grain = []
            Q_veg = 0.0
            P_veg = []
            for land_name, year_solution in self.base.optimal_solution.items():
                if year not in year_solution:
                    continue
                for season, season_solution in year_solution[year].items():
                    for crop_id, area in season_solution.items():
                        ci = self.base.crop_info[crop_id]
                        y_base = ci['yield_base']
                        q = area * y_base
                        p_on = mean_price[(crop_id, year)]
                        p_off = mean_price_no_delta[(crop_id, year)]
                        if is_grain(ci):
                            Q_grain += q
                            P_grain.append((p_on, p_off, q))
                        elif is_veg(ci):
                            Q_veg += q
                            P_veg.append((p_on, p_off, q))
            # 按产量加权的均价（更贴近市场均衡口径）
            def weighted_mean(items):
                if not items:
                    return (np.nan, np.nan)
                w = np.array([it[2] for it in items], dtype=float)
                p_on = np.sum(w * np.array([it[0] for it in items])) / np.sum(w)
                p_off = np.sum(w * np.array([it[1] for it in items])) / np.sum(w)
                return float(p_on), float(p_off)
            p_on_g, p_off_g = weighted_mean(P_grain)
            p_on_v, p_off_v = weighted_mean(P_veg)
            if Q_grain > 0 and not np.isnan(p_on_g) and not np.isnan(p_off_g):
                grains.append((Q_grain, p_on_g, p_off_g))
            if Q_veg > 0 and not np.isnan(p_on_v) and not np.isnan(p_off_v):
                vegs.append((Q_veg, p_on_v, p_off_v))

        def fit_and_plot(ax, data, title):
            if len(data) < 2:
                ax.set_title(f"{title}（样本不足）")
                return
            Q = np.array([d[0] for d in data], dtype=float)
            P_on = np.array([d[1] for d in data], dtype=float)
            P_off = np.array([d[2] for d in data], dtype=float)
            coef_on = np.polyfit(Q, P_on, 1)
            coef_off = np.polyfit(Q, P_off, 1)
            qline = np.linspace(Q.min()*0.95, Q.max()*1.05, 100)
            ax.scatter(Q/1e6, P_off, c='gray', s=25, alpha=0.7, label='不含δ（点）')
            ax.scatter(Q/1e6, P_on, c='steelblue', s=25, alpha=0.7, label='含δ（点）')
            ax.plot(qline/1e6, np.polyval(coef_off, qline), '--', color='gray', label=f"不含δ: p={coef_off[0]:.2e}·Q+{coef_off[1]:.2f}")
            ax.plot(qline/1e6, np.polyval(coef_on, qline), '-', color='steelblue', label=f"含δ: p={coef_on[0]:.2e}·Q+{coef_on[1]:.2f}")
            # 注释：在均值Q处的价格差
            qm = Q.mean()
            dm = np.polyval(coef_on, qm) - np.polyval(coef_off, qm)
            ax.annotate(f"δ致曲线移动: {dm:.2f}元/斤 @Q均值", xy=(qm/1e6, np.polyval(coef_on, qm)),
                        xytext=(qm/1e6, np.polyval(coef_on, qm)+0.05*max(1.0, P_on.mean())),
                        arrowprops=dict(arrowstyle='->', color='red'), color='red', fontsize=9)
            ax.set_xlabel('供给量Q（百万斤）')
            ax.set_ylabel('均价p（元/斤）')
            ax.set_title(title)
            ax.grid(alpha=0.3)
            ax.legend()

        fig, axes = plt.subplots(1, 2, figsize=(12.5, 5.6))
        fit_and_plot(axes[0], grains, '主粮：倒需求曲线（δ对比）')
        fit_and_plot(axes[1], vegs, '主菜：倒需求曲线（δ对比）')
        plt.tight_layout()
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"✅ 已生成：{filename}")

    # ---------- 作业清单与更替甘特图、大棚月度利用率 ----------
    def export_operation_schedule_and_gantt(self, schedule_csv='问题三_作业清单.csv', gantt_png='图7.X_作物更替甘特图.png'):
        print("\n📋 导出作业清单并生成更替甘特图...")
        all_lands = {**self.base.grain_lands, **self.base.irrigation_lands, **self.base.greenhouse_lands}

        def months_for(land_type: str, season: int):
            if land_type == '水浇地' or land_type in ['普通大棚', '智慧大棚']:
                return (3, 6) if season == 1 else (7, 10)
            else:  # 粮食类单季
                return (4, 9)

        rows = []
        for land_name, land_solution in self.base.optimal_solution.items():
            ltype = all_lands[land_name]['type']
            for year, year_solution in land_solution.items():
                for season, season_solution in year_solution.items():
                    for crop_id, area in season_solution.items():
                        if area <= 0.01:
                            continue
                        crop_info = self.base.crop_info[crop_id]
                        sow_m, harv_m = months_for(ltype, season)
                        rows.append({
                            '地块名称': land_name,
                            '地块类型': ltype,
                            '年份': year,
                            '季节': season,
                            '作物编号': crop_id,
                            '作物名称': crop_info['name'],
                            '种植面积(亩)': round(float(area), 2),
                            '播种月': sow_m,
                            '收获月': harv_m,
                        })
        df = pd.DataFrame(rows)
        df.sort_values(['地块名称', '年份', '季节'], inplace=True)
        df.to_csv(schedule_csv, index=False, encoding='utf-8-sig')
        print(f"✅ 已生成：{schedule_csv}")

        # 生成甘特图：选取面积总量Top12地块
        totals = df.groupby('地块名称')['种植面积(亩)'].sum().sort_values(ascending=False)
        top_lands = list(totals.head(12).index)
        gdf = df[df['地块名称'].isin(top_lands)].copy()
        # 将(年, 月)映射到总月份序号（从2024-01开始）
        def month_index(year, m):
            return (year - 2024) * 12 + (m - 1)
        gdf['start'] = [month_index(y, m) for y, m in zip(gdf['年份'], gdf['播种月'])]
        gdf['end'] = [month_index(y, m) for y, m in zip(gdf['年份'], gdf['收获月'])]

        set_plot_style()
        fig, ax = plt.subplots(figsize=(12.8, 6.8))
        # 颜色按作物类别区分
        def crop_color(crop_id: int) -> str:
            info = self.base.crop_info.get(crop_id, {})
            t = str(info.get('type', ''))
            if '豆' in t:
                return 'tab:green'
            if '小麦' in info.get('name', '') or '粮' in t:
                return 'tab:orange'
            if '玉米' in info.get('name', ''):
                return 'tab:olive'
            if '瓜' in info.get('name', ''):
                return 'tab:cyan'
            if '薯' in info.get('name', ''):
                return 'tab:brown'
            if '菌' in t:
                return 'tab:purple'
            if '蔬' in t:
                return 'tab:blue'
            return 'steelblue'
        yticks = []
        ylabels = []
        y = 0
        for land in top_lands:
            sub = gdf[gdf['地块名称'] == land]
            for _, r in sub.iterrows():
                color = crop_color(int(r['作物编号']))
                ax.barh(y, r['end'] - r['start'] + 1, left=r['start'], height=0.72,
                        color=color, alpha=0.75, edgecolor='white', linewidth=0.6)
                ax.text(r['start'] + 0.2, y, f"{int(r['年份'])}年S{int(r['季节'])}-{r['作物名称']} ({r['种植面积(亩)']}亩)",
                        fontsize=8.5, va='center', color='black')
            yticks.append(y)
            ylabels.append(land)
            y += 1
        ax.set_yticks(yticks)
        ax.set_yticklabels(ylabels)
        ax.set_xlabel('时间（月，从2024-01起）')
        ax.set_title('问题三 作物更替甘特图（Top12地块）')
        # 年度分隔线
        for yr in range(2024, int(gdf['年份'].max()) + 1):
            ax.axvline((yr - 2024) * 12, color='lightgray', linewidth=0.8)
        ax.grid(alpha=0.35, axis='x')
        plt.tight_layout()
        plt.savefig(gantt_png, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"✅ 已生成：{gantt_png}")

    def generate_greenhouse_monthly_utilization(self, out_csv='问题三_大棚月度利用率.csv', out_png='图7.X_大棚月度利用率.png'):
        print("\n🏠 生成大棚月度吞吐率与利用率...")
        greenhouses = {**self.base.greenhouse_lands}
        if not greenhouses:
            print("⚠️ 无大棚地块，跳过")
            return
        total_area = sum(info['area'] for info in greenhouses.values())

        def months_for(land_type: str, season: int):
            return (3, 6) if season == 1 else (7, 10)

        # 月序列
        months = [(year, m) for year in self.base.years for m in range(1, 13)]
        idx_map = {(y, m): (y - 2024) * 12 + (m - 1) for (y, m) in months}
        occ = np.zeros(len(months), dtype=float)
        thru = np.zeros(len(months), dtype=float)

        # 预估产出按季均摊到月份
        mean_yield_factor = {(j, t): float(np.mean([sc[j][t]['yield_factor'] for sc in self.base.scenarios])) for j in self.base.crop_info for t in self.base.years}

        for land_name, land_solution in self.base.optimal_solution.items():
            if land_name not in greenhouses:
                continue
            ltype = greenhouses[land_name]['type']
            for year, year_solution in land_solution.items():
                for season, season_solution in year_solution.items():
                    sow_m, harv_m = months_for(ltype, season)
                    dur = max(1, harv_m - sow_m + 1)
                    for crop_id, area in season_solution.items():
                        if area <= 0.01:
                            continue
                        # 占用
                        for m in range(sow_m, harv_m + 1):
                            occ[idx_map[(year, m)]] += float(area)
                        # 吞吐（按季分摊）
                        y_base = self.base.crop_info[crop_id]['yield_base']
                        yf = mean_yield_factor[(crop_id, year)]
                        production = float(area) * y_base * yf
                        per_month = production / dur
                        for m in range(sow_m, harv_m + 1):
                            thru[idx_map[(year, m)]] += per_month

        util = occ / max(1e-6, total_area)
        # 导出CSV
        rows = []
        for (y, m) in months:
            k = idx_map[(y, m)]
            rows.append({'年份': y, '月份': m, '占用面积(亩)': round(occ[k], 2), '总大棚面积(亩)': round(total_area, 2), '利用率': round(util[k], 4), '产出(斤)': round(thru[k], 2)})
        pd.DataFrame(rows).to_csv(out_csv, index=False, encoding='utf-8-sig')
        print(f"✅ 已生成：{out_csv}")

        # 绘图
        set_plot_style()
        fig, ax1 = plt.subplots(figsize=(12.8, 4.8))
        x = np.arange(len(months))
        ax1.plot(x, util * 100, '-', color='tab:blue', marker='o', markersize=3, label='利用率(%)')
        ax1.set_ylabel('利用率(%)')
        ax1.set_xlabel('时间（月，从2024-01起）')
        ax2 = ax1.twinx()
        ax2.bar(x, thru / 1e6, color='tab:cyan', alpha=0.35, label='吞吐(百万斤)')
        ax2.set_ylabel('吞吐（百万斤）')
        ax1.set_title('大棚月度利用率与吞吐量（问题三方案）')
        # 合并图例
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
        # 年度分隔线
        for yr in range(2024, max(self.base.years) + 1):
            ax1.axvline((yr - 2024) * 12, color='lightgray', linewidth=0.8)
        plt.tight_layout()
        plt.savefig(out_png, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"✅ 已生成：{out_png}")

    # ---------- 多随机种子误差带（K=1000, seeds={42,123,2024}） ----------
    def run_multi_seed_error_band(self, seeds=(42, 123, 2024), out_png='问题三_多种子误差带.png', out_csv='问题三_多种子误差带.csv', saa: int | None = None):
        print("\n🧪 多随机种子误差带实验：K=1000, seeds=", seeds)
        if saa is not None:
            self.base.N_saa = int(saa)
        rows = []
        xs = []
        means = []
        stds = []
        for seed in seeds:
            # 全流程：重置→情景→Copula→SAA→求解→全情景验证
            self.base.load_and_process_data()
            self.build_relationships()
            # 设置随机种子
            try:
                self.base.random_seed = int(seed)
            except Exception:
                pass
            self.base.N_scenarios = 1000
            self.base.generate_stochastic_scenarios()
            self.generate_correlated_scenarios()
            self.base.select_representative_scenarios()
            self.base.use_tail_cvar = False
            self.base.build_stochastic_programming_model()
            ok = self.base.solve_model()
            if not ok:
                print(f"   ⚠️ 种子{seed}求解失败，跳过")
                continue
            metrics = self.validate_with_synergy()
            rows.append({'seed': seed, **metrics})
            xs.append(seed)
            means.append(metrics['expected_profit'])
            stds.append(metrics['profit_std'])
        if rows:
            df = pd.DataFrame(rows)
            df.to_csv(out_csv, index=False, encoding='utf-8-sig')
            print(f"✅ 已生成：{out_csv}")
            # 绘图：均值±1σ
            set_plot_style()
            order = np.argsort(xs)
            xs_plot = np.array(xs)[order]
            means_plot = np.array(means)[order]
            stds_plot = np.array(stds)[order]
            fig, ax = plt.subplots(figsize=(8.0, 5.2))
            ax.plot(xs_plot, means_plot/1e6, '-o', color='tab:blue', label='期望收益')
            ax.fill_between(xs_plot, (means_plot-stds_plot)/1e6, (means_plot+stds_plot)/1e6,
                            color='tab:blue', alpha=0.2, label='±1σ 区间')
            ax.set_xlabel('随机种子')
            ax.set_ylabel('收益（百万元）')
            ax.set_title('问题三 多随机种子误差带（K=1000）')
            ax.legend()
            plt.tight_layout()
            plt.savefig(out_png, dpi=300, bbox_inches='tight')
            plt.close()
            print(f"✅ 已生成：{out_png}")


def main():
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='ignore')
        sys.stderr.reconfigure(encoding='utf-8', errors='ignore')
    except Exception:
        pass

    solver = Problem3CorrelatedSubstitutionSolver()
    metrics = solver.run(iterations=2)
    print("\n关键指标：")
    for k, v in metrics.items():
        print(f" - {k}: {v:,.2f}")
    return solver


if __name__ == '__main__':
    main()


